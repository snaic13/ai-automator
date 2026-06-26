import io
import requests
from bs4 import BeautifulSoup
from PyPDF2 import PdfReader
from docx import Document
from openpyxl import load_workbook
from openai import OpenAI
from config import OPENAI_API_KEY, MODEL, MAX_TOKENS

client = OpenAI(api_key=OPENAI_API_KEY, timeout=120.0)

MAX_INPUT_CHARS = 12000


def truncate(text: str) -> str:
    if len(text) > MAX_INPUT_CHARS:
        return text[:MAX_INPUT_CHARS] + "\n\n[текст обрезан]"
    return text


def extract_pdf_text(pdf_bytes: bytes) -> str:
    try:
        reader = PdfReader(io.BytesIO(pdf_bytes))
        text = ""
        for page in reader.pages:
            t = page.extract_text()
            if t:
                text += t + "\n"
        return text.strip() if text.strip() else "Не удалось извлечь текст из PDF"
    except Exception as e:
        return f"Ошибка чтения PDF: {e}"


def extract_docx_text(docx_bytes: bytes) -> str:
    try:
        doc = Document(io.BytesIO(docx_bytes))
        text = "\n".join([p.text for p in doc.paragraphs if p.text.strip()])
        return text.strip() if text.strip() else "Документ пуст"
    except Exception as e:
        return f"Ошибка чтения Word: {e}"


def extract_xlsx_text(xlsx_bytes: bytes) -> str:
    try:
        wb = load_workbook(io.BytesIO(xlsx_bytes), read_only=True, data_only=True)
        text = ""
        for sheet in wb.sheetnames:
            ws = wb[sheet]
            text += f"--- Лист: {sheet} ---\n"
            for row in ws.iter_rows(values_only=True):
                row_text = [str(c) if c is not None else "" for c in row]
                line = " | ".join(row_text)
                if line.strip(" |"):
                    text += line + "\n"
            text += "\n"
        wb.close()
        return text.strip() if text.strip() else "Таблица пуста"
    except Exception as e:
        return f"Ошибка чтения Excel: {e}"


def extract_csv_text(csv_bytes: bytes) -> str:
    try:
        return csv_bytes.decode("utf-8", errors="replace")
    except Exception as e:
        return f"Ошибка чтения CSV: {e}"


def automate(prompt: str, system: str = "Ты полезный AI-ассистент для бизнес-автоматизации.") -> str:
    response = client.chat.completions.create(
        model=MODEL,
        messages=[
            {"role": "system", "content": system},
            {"role": "user", "content": prompt},
        ],
        max_tokens=MAX_TOKENS,
    )
    return response.choices[0].message.content


def process_document(text: str) -> str:
    return automate(
        f"Обработай документ и верни структурированный ответ:\n\n{truncate(text)}",
        "Ты эксперт по обработке документов. Извлеки ключевую информацию, сумму, даты, стороны.",
    )


def customer_support(question: str) -> str:
    return automate(
        truncate(question),
        "Ты сотрудник поддержки клиентов. Отвечай вежливо, professionally, на русском языке.",
    )


def generate_report(data: str) -> str:
    return automate(
        f"Сгенерируй отчёт по данным:\n\n{truncate(data)}",
        "Ты аналитик. Создай краткий отчёт с выводами и рекомендациями.",
    )


def summarize(text: str) -> str:
    return automate(
        f"Сделай краткое резюме:\n\n{truncate(text)}",
        "Ты редактор. Создай сжатое резюме, выдели ключевые моменты.",
    )


def translate(text: str) -> str:
    return automate(
        f"Переведи текст на английский язык:\n\n{truncate(text)}",
        "Ты профессиональный переводчик. Переведи точно, сохраняя смысл.",
    )


def extract_emails(text: str) -> str:
    return automate(
        f"Извлеки все email-адреса из текста:\n\n{truncate(text)}",
        "Ты парсер. Извлеки только email-адреса, верни списком.",
    )


def customer_support(question: str) -> str:
    return automate(
        question,
        "Ты сотрудник поддержки клиентов. Отвечай вежливо, professionally, на русском языке.",
    )


def generate_report(data: str) -> str:
    return automate(
        f"Сгенерируй отчёт по данным:\n\n{data}",
        "Ты аналитик. Создай краткий отчёт с выводами и рекомендациями.",
    )


def summarize(text: str) -> str:
    return automate(
        f"Сделай краткое резюме:\n\n{text}",
        "Ты редактор. Создай сжатое резюме, выдели ключевые моменты.",
    )


def translate(text: str) -> str:
    return automate(
        f"Переведи текст на английский язык:\n\n{text}",
        "Ты профессиональный переводчик. Переведи точно, сохраняя смысл.",
    )


def extract_emails(text: str) -> str:
    return automate(
        f"Извлеки все email-адреса из текста:\n\n{text}",
        "Ты парсер. Извлеки только email-адреса, верни списком.",
    )


def fetch_url(url: str) -> str:
    try:
        headers = {"User-Agent": "Mozilla/5.0"}
        resp = requests.get(url, headers=headers, timeout=15, allow_redirects=True)
        resp.raise_for_status()
        soup = BeautifulSoup(resp.text, "html.parser")
        for tag in soup(["script", "style", "nav", "footer", "header", "iframe"]):
            tag.decompose()
        text = soup.get_text(separator="\n", strip=True)
        lines = [l.strip() for l in text.splitlines() if l.strip()]
        text = "\n".join(lines[:300])
        if len(text) > 8000:
            text = text[:8000]
        return text
    except Exception as e:
        return f"Ошибка при загрузке URL: {e}"


def summarize_url(url: str) -> str:
    content = fetch_url(url)
    if content.startswith("Ошибка"):
        return content
    return automate(
        f"Сделай краткое резюме содержимого страницы:\n\n{content[:6000]}",
        "Ты аналитик. Создай структурированное резюме основных тезисов.",
    )


def process_image_text(text: str) -> str:
    return automate(
        f"Проанализируй текст из изображения и структурируй его:\n\n{text}",
        "Ты помощник. Определи тип документа, извлеки ключевую информацию.",
    )


chat_history = []


def chat(message: str) -> str:
    chat_history.append({"role": "user", "content": message})
    if len(chat_history) > 20:
        chat_history.pop(0)
        chat_history.pop(0)
    try:
        messages = [{"role": "system", "content": "Ты умный AI-ассистент. Отвечай на русском языке, подробно и по делу. Если вопрос требует кода — покажи его. Если творческий — будь креативным."}]
        messages.extend(chat_history)
        response = client.chat.completions.create(
            model=MODEL,
            messages=messages,
            max_tokens=MAX_TOKENS,
        )
        reply = response.choices[0].message.content
        chat_history.append({"role": "assistant", "content": reply})
        return reply
    except Exception as e:
        return f"Ошибка: {e}"


def generate_image_idea(prompt: str) -> str:
    return automate(
        f"Создай детальный промпт для генерации изображения нейросетью (Midjourney/DALL-E/Stable Diffusion) по описанию:\n\n{truncate(prompt)}",
        "Ты эксперт по промпт-инжинирингу. Создай детальный промпт на английском языке с описанием стиля, освещения, композиции, деталей.",
    )


def business_idea(prompt: str) -> str:
    return automate(
        f"Предложи бизнес-идею и план по описанию:\n\n{truncate(prompt)}",
        "Ты бизнес-консультант. Предложи конкретную бизнес-идею с анализом рынка, конкурентами, источниками дохода и первыми шагами.",
    )


def resume_improve(prompt: str) -> str:
    return automate(
        f"Улучши резюме или опиши опыт для резюме:\n\n{truncate(prompt)}",
        "Ты HR-эксперт. Улучши формулировки, добавь достижения с цифрами, сделай профессиональным и привлекательным для работодателя.",
    )


def legal_review(prompt: str) -> str:
    return automate(
        f"Проанализируй юридический текст или договор:\n\n{truncate(prompt)}",
        "Ты юрист. Проанализируй текст, выдели ключевые условия, риски, сроки и обязательства. Дай рекомендации.",
    )


def math_solve(prompt: str) -> str:
    return automate(
        f"Реши задачу покажи решение пошагово:\n\n{truncate(prompt)}",
        "Ты математик. Реши задачу пошагово, покажи все вычисления и объясни логику.",
    )


def code_generate(prompt: str) -> str:
    return automate(
        f"Сгенерируй код по описанию:\n\n{truncate(prompt)}",
        "Ты программист. Сгенерируй чистый, рабочий код с комментариями на русском языке. Укажи язык программирования.",
    )


def email_compose(prompt: str) -> str:
    return automate(
        f"Напиши письмо по описанию:\n\n{truncate(prompt)}",
        "Ты бизнес-письменник. Напиши профессиональное, вежливое и структурированное письмо.",
    )


def social_post(prompt: str) -> str:
    return automate(
        f"Создай пост для соцсетей по описанию:\n\n{truncate(prompt)}",
        "Ты SMM-специалист. Создай engaging пост с хэштегами, эмодзи и призывом к действию. Адаптируй под платформу.",
    )


if __name__ == "__main__":
    print("=== AI-Automator MVP ===\n")

    # Тест 1: Обработка документа
    print("1. Обработка документа:")
    doc = "Счёт №123 от 25.06.2026. Поставщик: ООО Ромашка. Сумма: 150,000 руб."
    print(process_document(doc))
    print()

    # Тест 2: Поддержка клиентов
    print("2. Поддержка клиентов:")
    print(customer_support("Как отменить подписку?"))
    print()

    # Тест 3: Генерация отчёта
    print("3. Генерация отчёта:")
    report_data = "Продажи: Янв 100k, Фев 120k, Мар 95k, Апр 140k"
    print(generate_report(report_data))
